"""
Imaris data merger
------------------
Scrapes Imaris .xls exports and populates an overview workbook with
two sheets: IMARIS RAW DENDRITES and IMARIS RAW SPINES.

Filename format expected:
    <name> <number> <roi>.xls
    e.g.  "trine 1 oriens.xls", "kiara 3 radiatum.xls"

  - Name      = text before the first number (any capitalisation)
  - Number    = the first integer in the filename
  - ROI       = whatever text follows the number (before the extension)

Usage:
    python merge_imaris.py <output.xlsx> <folder_of_xls_files>

Behaviour:
  - If <output.xlsx> doesn't exist, a fresh overview is created.
  - If it does exist, new dendrites are appended; any (Name, Number)
    pair already present in a sheet is skipped for that sheet.
  - A detailed summary of successes and problems is printed at the end.
"""

from __future__ import annotations

import argparse
import difflib
import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
import xlrd  # noqa: F401  -- required by pandas to read .xls
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Sheet headers (keep in sync with the existing overview template)
# ---------------------------------------------------------------------------

DENDRITE_HEADERS = [
    "Animal ID", "Name (blinded)", "Dendrite number", "ROI (RAD/ORS)",
    "Slide number",
    "Dendrite Diameter Threshold", "Spine Seed Point Diameter (um)",
    "Spine Maximum Length (um)", "Spine Seed Point Threshold",
    "Spine Diameter Threshold",
    "Spine Diameter Algorithm (Distance Map / Cross Section)",
    "Variable", "Min", "Max", "Mean", "StdDev", "Median", "Sum", "Count",
    "Unit", "Category", "Collection", "Depth", "Distance", "Level", "Radius",
    "Time", "Type",
]

SPINE_HEADERS = [
    "Animal ID", "Name (blinded)", "Dendrite number", "ROI (RAD/ORS)",
    "Variable", "Min", "Max", "Mean", "StdDev", "Median", "Sum", "Count",
    "Unit", "Category", "Collection", "Depth", "Distance", "Level", "Radius",
    "Time", "Type",
]

# Column indexes (1-indexed) where numeric data lives -- used to apply the
# Imaris light-green fill and 2-decimal number format.
DENDRITE_NUMERIC_COLS = range(13, 20)   # Min..Count
SPINE_NUMERIC_COLS = range(6, 13)       # Min..Count


# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------

@dataclass
class ParsedFilename:
    name: str          # e.g. "Trine"
    dno: str           # zero-padded, e.g. "001"
    roi: str           # e.g. "Oriens"


# Filename regex: capture the leading text, the first integer, and whatever
# follows up to the extension. Whitespace, underscores and hyphens are all
# treated as separators between the three parts.
_FILENAME_RE = re.compile(
    r"""
    ^\s*
    (?P<name>[^\d]+?)             # name (no digits)
    [\s_\-]+
    (?P<num>\d+)                  # dendrite number
    [\s_\-]+
    (?P<roi>[A-Za-z][\w\s\-]*?)   # roi (letters, optional extras)
    \s*$
    """,
    re.VERBOSE,
)


def parse_filename(filename: str) -> ParsedFilename:
    stem = Path(filename).stem
    match = _FILENAME_RE.match(stem)
    if not match:
        raise ValueError(
            f"Filename '{filename}' does not match '<name> <number> <roi>' "
            f"(e.g. 'trine 1 oriens.xls')."
        )
    name_raw = match.group("name").strip().replace("_", " ").replace("-", " ")
    # Collapse whitespace and title-case the name so "mary poppins" -> "Mary Poppins"
    name = " ".join(w.capitalize() for w in name_raw.split())
    dno = match.group("num").zfill(3)
    roi_raw = match.group("roi").strip().replace("_", " ").replace("-", " ")
    roi = " ".join(w.capitalize() for w in roi_raw.split())
    return ParsedFilename(name=name, dno=dno, roi=roi)


# ---------------------------------------------------------------------------
# Sheet lookup (tolerant of case, whitespace, small typos)
# ---------------------------------------------------------------------------

def _find_sheet(xls_path: Path, target: str) -> str:
    xl = pd.ExcelFile(xls_path)
    target_low = target.lower().strip()
    for name in xl.sheet_names:
        if name.lower().strip() == target_low:
            return name
    for name in xl.sheet_names:
        if target_low in name.lower().strip():
            return name
    candidates = [n.lower().strip() for n in xl.sheet_names]
    matches = difflib.get_close_matches(target_low, candidates, n=1, cutoff=0.7)
    if matches:
        idx = candidates.index(matches[0])
        return xl.sheet_names[idx]
    raise ValueError(
        f"no sheet matching '{target}' (found: {xl.sheet_names})"
    )


# ---------------------------------------------------------------------------
# Algorithm-sheet parsing
# ---------------------------------------------------------------------------

ALGO_KEYS = {
    "Dendrite Diameter Threshold": "Dendrite Diameter Threshold",
    "Spine Seed Point Diameter (um)": "Spine Seed Point Diameter",
    "Spine Maximum Length (um)": "Spine Maximum Length",
    "Spine Seed Point Threshold": "Spine Seed Point Threshold",
    "Spine Diameter Threshold": "Spine Diameter Threshold",
    "Spine Diameter Algorithm (Distance Map / Cross Section)": "Spine Diameter Algorithm",
}


def parse_algorithm(df_algo: pd.DataFrame) -> dict[str, object]:
    mapping: dict[str, object] = {}
    for raw in df_algo.iloc[:, 0].dropna().astype(str):
        if "=" not in raw:
            continue
        k, v = raw.split("=", 1)
        k = k.strip()
        v = v.strip().removesuffix(" um").strip()
        try:
            mapping[k] = float(v)
        except ValueError:
            mapping[k] = v
    return mapping


def algo_values(xls_path: Path) -> dict[str, object]:
    sheet = _find_sheet(xls_path, "Algorithm")
    df_algo = pd.read_excel(xls_path, sheet_name=sheet, header=None)
    parsed = parse_algorithm(df_algo)
    return {col: parsed.get(key, "") for col, key in ALGO_KEYS.items()}


# ---------------------------------------------------------------------------
# Average / Spines sheet reading
# ---------------------------------------------------------------------------

AVERAGE_COLS = [
    "Variable", "Min", "Max", "Mean", "StdDev", "Median", "Sum", "Count",
    "Unit", "Category", "Collection", "Depth", "Distance", "Level", "Radius",
    "Time", "Type",
]

SPINE_COLS = [
    "Variable", "Min", "Max", "Mean", "StdDev", "Median", "Sum", "Count",
    "Unit", "Category", "Collection", "Depth", "Distance", "Level", "Radius",
    "Surpass Object",  # -> overview "Time" column (spine type: Stubby/Mushroom/...)
    "Time",            # -> overview "Type" column (constant 1)
]


def read_average(xls_path: Path) -> pd.DataFrame:
    sheet = _find_sheet(xls_path, "Average")
    df = pd.read_excel(xls_path, sheet_name=sheet, header=1)
    for c in AVERAGE_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[AVERAGE_COLS].copy()


def read_spines(xls_path: Path) -> pd.DataFrame:
    sheet = _find_sheet(xls_path, "Spines")
    df = pd.read_excel(xls_path, sheet_name=sheet, header=1)
    for c in SPINE_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[SPINE_COLS].copy()


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------

def build_dendrite_rows(xls_path: Path, pf: ParsedFilename) -> list[list[object]]:
    avg_df = read_average(xls_path)
    try:
        algo = algo_values(xls_path)
    except ValueError:
        algo = {col: "" for col in ALGO_KEYS}

    rows: list[list[object]] = []
    for i, row in enumerate(avg_df.itertuples(index=False)):
        first = i == 0
        out: list[object] = [
            "",           # Animal ID
            pf.name,      # Name (blinded)
            pf.dno,       # Dendrite number
            pf.roi,       # ROI
            "",           # Slide number
            algo["Dendrite Diameter Threshold"] if first else "",
            algo["Spine Seed Point Diameter (um)"] if first else "",
            algo["Spine Maximum Length (um)"] if first else "",
            algo["Spine Seed Point Threshold"] if first else "",
            algo["Spine Diameter Threshold"] if first else "",
            algo["Spine Diameter Algorithm (Distance Map / Cross Section)"] if first else "",
        ]
        out.extend(list(row))
        rows.append(out)
    return rows


def build_spine_rows(xls_path: Path, pf: ParsedFilename) -> list[list[object]]:
    spines_df = read_spines(xls_path)
    rows: list[list[object]] = []
    for row in spines_df.itertuples(index=False):
        out: list[object] = ["", pf.name, pf.dno, pf.roi]
        out.extend(list(row))
        rows.append(out)
    return rows


# ---------------------------------------------------------------------------
# Workbook creation / loading
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True)
IMARIS_FILL = PatternFill("solid", fgColor="CCFFCC")
IMARIS_NUM_FORMAT = "0.00"


def _init_sheet(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
    # Reasonable default column widths so the header is readable.
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 2)


def create_fresh_workbook(path: Path) -> None:
    wb = Workbook()
    ws_dend = wb.active
    ws_dend.title = "IMARIS RAW DENDRITES"
    _init_sheet(ws_dend, DENDRITE_HEADERS)

    ws_spine = wb.create_sheet("IMARIS RAW SPINES")
    _init_sheet(ws_spine, SPINE_HEADERS)

    wb.save(path)


# ---------------------------------------------------------------------------
# Overview helpers
# ---------------------------------------------------------------------------

def find_next_empty_row(ws) -> int:
    for row in range(2, ws.max_row + 2):
        if all(ws.cell(row=row, column=c).value in (None, "") for c in (2, 3)):
            return row
    return ws.max_row + 1


def existing_keys(ws) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        dno = ws.cell(row=row, column=3).value
        if name and dno:
            keys.add((str(name).strip(), str(dno).strip()))
    return keys


def write_rows(ws, start_row: int, rows: list[list[object]], numeric_cols: range) -> int:
    current = start_row
    for row_data in rows:
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current, column=col_idx, value=val)
            if col_idx in numeric_cols and isinstance(val, (int, float)):
                cell.fill = IMARIS_FILL
                cell.number_format = IMARIS_NUM_FORMAT
        current += 1
    return current


# ---------------------------------------------------------------------------
# Summary reporting
# ---------------------------------------------------------------------------

@dataclass
class RunSummary:
    total_files: int = 0
    dendrites_written: int = 0
    dendrite_rows: int = 0
    spines_written: int = 0
    spine_rows: int = 0
    dendrites_skipped_duplicate: list[str] = field(default_factory=list)
    spines_skipped_duplicate: list[str] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)

    def print(self) -> None:
        print()
        print("=" * 60)
        print("SUMMARY")
        print("=" * 60)
        print(f"Files processed:            {self.total_files}")
        print(f"Dendrites written (files):  {self.dendrites_written}  ({self.dendrite_rows} rows)")
        print(f"Spines written (files):     {self.spines_written}  ({self.spine_rows} rows)")
        if self.dendrites_skipped_duplicate:
            print(f"Dendrites skipped as already-present: {len(self.dendrites_skipped_duplicate)}")
            for s in self.dendrites_skipped_duplicate:
                print(f"  - {s}")
        if self.spines_skipped_duplicate:
            print(f"Spines skipped as already-present:    {len(self.spines_skipped_duplicate)}")
            for s in self.spines_skipped_duplicate:
                print(f"  - {s}")
        if self.problems:
            print()
            print(f"PROBLEMS ({len(self.problems)}):")
            for p in self.problems:
                print(f"  - {p}")
        else:
            print()
            print("No problems encountered.")
        print("=" * 60)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def merge(output_path: Path, input_folder: Path) -> None:
    if not output_path.exists():
        print(f"Creating fresh overview at {output_path}")
        create_fresh_workbook(output_path)
    else:
        print(f"Using existing overview at {output_path}")

    xls_files = sorted(input_folder.glob("*.xls")) + sorted(input_folder.glob("*.xlsx"))
    if not xls_files:
        raise SystemExit(f"No .xls/.xlsx files found in {input_folder}")

    wb = load_workbook(output_path)

    if "IMARIS RAW DENDRITES" not in wb.sheetnames or "IMARIS RAW SPINES" not in wb.sheetnames:
        raise SystemExit(
            "Output file is missing required sheets. "
            "Delete the output file and let the script re-create it."
        )

    ws_dend = wb["IMARIS RAW DENDRITES"]
    ws_spine = wb["IMARIS RAW SPINES"]

    dend_keys = existing_keys(ws_dend)
    spine_keys = existing_keys(ws_spine)
    dend_row = find_next_empty_row(ws_dend)
    spine_row = find_next_empty_row(ws_spine)

    summary = RunSummary(total_files=len(xls_files))

    for xls_path in xls_files:
        print(f"\n{xls_path.name}")
        try:
            pf = parse_filename(xls_path.name)
        except Exception as e:
            print(f"  SKIPPED: {e}")
            summary.problems.append(f"{xls_path.name}: filename unparseable -- {e}")
            continue

        label = f"{pf.name} {pf.dno} {pf.roi}"
        print(f"  parsed as: {label}")

        # -- Dendrites --
        if (pf.name, pf.dno) in dend_keys:
            print("  dendrites: already present, skipped")
            summary.dendrites_skipped_duplicate.append(f"{label} ({xls_path.name})")
        else:
            try:
                rows = build_dendrite_rows(xls_path, pf)
                dend_row = write_rows(ws_dend, dend_row, rows, DENDRITE_NUMERIC_COLS)
                dend_keys.add((pf.name, pf.dno))
                summary.dendrites_written += 1
                summary.dendrite_rows += len(rows)
                print(f"  dendrites: wrote {len(rows)} rows")
            except Exception as e:
                print(f"  dendrites: FAILED -- {e}")
                summary.problems.append(f"{label}: dendrite data failed -- {e}")

        # -- Spines --
        if (pf.name, pf.dno) in spine_keys:
            print("  spines:    already present, skipped")
            summary.spines_skipped_duplicate.append(f"{label} ({xls_path.name})")
        else:
            try:
                rows = build_spine_rows(xls_path, pf)
                spine_row = write_rows(ws_spine, spine_row, rows, SPINE_NUMERIC_COLS)
                spine_keys.add((pf.name, pf.dno))
                summary.spines_written += 1
                summary.spine_rows += len(rows)
                print(f"  spines:    wrote {len(rows)} rows")
            except Exception as e:
                print(f"  spines:    FAILED -- {e}")
                summary.problems.append(f"{label}: spine data failed -- {e}")

    wb.save(output_path)
    print(f"\nSaved to {output_path}")
    summary.print()


def cli() -> None:
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("output", type=Path,
                   help="Output .xlsx file. Created fresh if it doesn't exist; "
                        "appended to (skipping duplicates) if it does.")
    p.add_argument("input_folder", type=Path,
                   help="Folder containing Imaris .xls files named "
                        "'<name> <number> <roi>.xls'")
    args = p.parse_args()
    merge(args.output, args.input_folder)


if __name__ == "__main__":
    cli()
